#!D:\Users\Anaconda\anaconda3\python.exe
# -*- coding: utf-8 -*-
# @Time    : 2026/1/23 18:18
# @Author  : Coffee
# @Project : Auxiliary
# @File    : ExcelPlot.py

from openpyxl import Workbook
from openpyxl.chart import *
from openpyxl.chart.label import DataLabelList
from openpyxl.reader.excel import load_workbook


SN = "2506210102"
dirr = 'C:/Users/w00025121/Desktop/SNP DATA.xlsx'
wb = load_workbook(dirr)
ws = wb[SN]
titles = ['Z-n', 'C-n', 'deltaC-n' ,'S11-n']



categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

for i in range(4):
  chart = LineChart()
  chart.title = titles[i]

  data = Reference(ws, min_col=i+2, min_row=2, max_row=ws.max_row)

  chart.add_data(data, titles_from_data=False)
  chart.set_categories(categories)

  chart.dataLabels = DataLabelList()

  ws.add_chart(chart, "E2")

wb.save(r"C:\Users\w00025121\Desktop\SNP DATA.xlsx")
